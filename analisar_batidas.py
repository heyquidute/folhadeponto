from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

def analisar_batidas(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)

    if "DIVERGENCIA" in wb.sheetnames:
        del wb["DIVERGENCIA"]
    
    aba_div = wb.create_sheet("DIVERGENCIA", 0)
    aba_div.sheet_view.showGridLines = False

    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    borda_inferior = Border(bottom=Side(style="thin", color="000000"))
    preenchimento_azul = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Cabeçalho
    cabecalho = ["Funcionário", "Data"]
    aba_div.append(cabecalho)
    for cel in aba_div[1]:
        cel.font = bold_font
        cel.alignment = center_align
        cel.border = borda_inferior

    # Lista funcionários com divergências
    funcionarios_com_div = set()

    for nome_aba in wb.sheetnames.copy():
        if nome_aba == "DIVERGENCIA":
            continue

        aba = wb[nome_aba]

        cabecalhos = [celula.value for celula in next(aba.iter_rows(min_row=1, max_row=1))]
        if "Data" not in cabecalhos or "Hr Tot T" not in cabecalhos:
            continue

        idx_data = cabecalhos.index("Data")+1
        idx_hrtot = cabecalhos.index("Hr Tot T")+1

        for linha_celulas in aba.iter_rows(min_row=2):
            data = linha_celulas[idx_data-1].value
            valor_hrtot = linha_celulas[idx_hrtot-1].value

            if not valor_hrtot:
                continue

            try:
                negativo = "-" in str(valor_hrtot) or valor_hrtot.startswith("−")
            except:
                negativo = False

            if negativo:
                for cel in linha_celulas:
                    cel.fill = preenchimento_azul

                aba_div.append([nome_aba,data])
                funcionarios_com_div.add(nome_aba)
    
    for nome_aba in wb.sheetnames.copy():
        if nome_aba != "DIVERGENCIA" and nome_aba not in funcionarios_com_div:
            del wb[nome_aba]

    for coluna in aba_div.columns:
        coluna_letra = coluna[0].column_letter
        max_len = max(len(str(cel.value)) if cel.value else 0 for cel in coluna)
        aba_div.column_dimensions[coluna_letra].width = max_len + 2

    for linha in aba_div.iter_rows(min_row=2):
        for cel in linha:
            cel.alignment = center_align

    wb.save(caminho_arquivo)
    print(f"Relatorio de divergencias salvo com sucesso:\n{caminho_arquivo}")