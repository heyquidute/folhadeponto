from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from str_to_time import str_para_tempo

def analisar_tudo(caminho_arquivo):
    # Carrega o workbook existente
    wb = load_workbook(caminho_arquivo)

    # Remove a aba "OCORRÊNCIAS" se já existir
    if "OCORRÊNCIAS" in wb.sheetnames:
        del wb["OCORRÊNCIAS"]

    # Cria nova aba OCORRÊNCIAS no início
    aba_ocorrencias = wb.create_sheet("OCORRÊNCIAS", 0)
    aba_ocorrencias.sheet_view.showGridLines = False

    # Estilos
    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    borda_inferior = Border(bottom=Side(style="thin", color="000000"))
    preenchimento_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    preenchimento_amarelo = PatternFill(start_color="F9E700", end_color="F9E700", fill_type="solid")
    preenchimento_laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    preenchimento_vermelho = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
    preenchimento_azul = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Cabeçalho da aba OCORRÊNCIAS
    cabecalho = ["Funcionário", "Data", "Ocorrência","Observações da folha de ponto"]
    aba_ocorrencias.append(cabecalho)
    for cel in aba_ocorrencias[1]:
        cel.font = bold_font
        cel.alignment = center_align
        cel.border = borda_inferior

    # Lista das abas dos funcionários com atestados
    funcionarios_com_atestado = set()

    # Percorre cada aba de funcionário
    for nome_aba in wb.sheetnames.copy():
        if nome_aba == "OCORRÊNCIAS":
            continue

        aba = wb[nome_aba]

        max_col = aba.max_column
        col_turno_manha = max_col - 3
        col_t_almoco = max_col - 2
        col_turno_tarde = max_col - 1
        col_total = max_col

        # Pega cabeçalhos da aba
        cabecalhos = [celula.value for celula in next(aba.iter_rows(min_row=1, max_row=1))]
        if "Data" not in cabecalhos or "Ocorrencia" not in cabecalhos or "Total" not in cabecalhos:
            continue

        idx_data = cabecalhos.index("Data") + 1
        idx_ocorrencia = cabecalhos.index("Ocorrencia") + 1
        idx_total = cabecalhos.index("Total") + 1
        idx_hrtot = cabecalhos.index("Hr Tot T")+1

        # Percorre as linhas da aba do funcionário
        for linha_celulas in aba.iter_rows(min_row=2):
            data = linha_celulas[idx_data - 1].value
            ocorrencia_raw = linha_celulas[idx_ocorrencia - 1].value
            total_raw = linha_celulas[idx_total - 1].value
            t_manha = str_para_tempo(linha_celulas[col_turno_manha - 1].value)
            t_almoco = str_para_tempo(linha_celulas[col_t_almoco - 1].value)
            t_tarde = str_para_tempo(linha_celulas[col_turno_tarde - 1].value)
            t_total = str_para_tempo(linha_celulas[col_total - 1].value)
            valor_hrtot = linha_celulas[idx_hrtot-1].value

            ocorrencia = str(ocorrencia_raw).strip().upper()

            # OCORRÊNCIAS
            if (ocorrencia.startswith("007") or ocorrencia.startswith("ATESTADO")):
                aba_ocorrencias.append([nome_aba, data, "Atestado médico", ocorrencia_raw])
                funcionarios_com_atestado.add(nome_aba)
                # Pinta a linha inteira de verde na aba do funcionário
                for cel in linha_celulas:
                    cel.fill = preenchimento_verde

            elif (ocorrencia.startswith("008") or ocorrencia.startswith("004")):
                aba_ocorrencias.append([nome_aba, data, "Banco de horas", ocorrencia_raw])
                funcionarios_com_atestado.add(nome_aba)
                # Pinta a linha inteira de verde na aba do funcionário
                for cel in linha_celulas:
                    cel.fill = preenchimento_verde

            elif ocorrencia.startswith("014"):
                horas = None
                if isinstance(total_raw, timedelta):
                    horas = total_raw.total_seconds() / 3600
                elif isinstance(total_raw, datetime):
                    horas = total_raw.hour + total_raw.minute / 60
                elif isinstance(total_raw, str):
                    try:
                        partes = total_raw.split(":")
                        horas = int(partes[0]) + int(partes[1]) / 60
                    except:
                        pass
                if horas is not None and horas < 5:
                    horas_compensar = 6 - horas
                    h, m = divmod(horas_compensar * 60, 60)
                    aba_ocorrencias.append([nome_aba, data, "Saída antecipada", f"{int(h):02d}h{int(m):02d}min a compensar"])
                    funcionarios_com_atestado.add(nome_aba)

                    # Pinta a linha de amarelo
                    for cel in linha_celulas:
                        cel.fill = preenchimento_amarelo

            # HORÁRIOS
            # Tempo do almoço menor que 1 hora
            if t_almoco and t_almoco < timedelta(hours=1):
                for cell in linha_celulas:
                    cell.fill = preenchimento_amarelo
                aba_ocorrencias.append([nome_aba, data, "Almoço < 1h", f"Tempo de almoço: {t_almoco}"])

            # Turno com mais de 6 horas sem intervalo
            if t_manha and t_manha > timedelta(hours=6):
                for cell in linha_celulas:
                    cell.fill = preenchimento_laranja
                aba_ocorrencias.append([nome_aba, data, "Turno da manhã > 6h", f"Duração do turno: {t_manha}"])

            if t_tarde and t_tarde > timedelta(hours=6):
                for cell in linha_celulas:
                    cell.fill = preenchimento_laranja
                aba_ocorrencias.append([nome_aba, data, "Turno da tarde > 6h", f"Duração do turno: {t_tarde}"])
            
            # Jornada total acima de 10 horas
            if t_total and t_total > timedelta(hours=10):
                for cell in linha_celulas:
                    cell.fill = preenchimento_vermelho
                aba_ocorrencias.append([nome_aba, data, "Jornada > 10h", f"Jornada total: {t_total}"])

            #ERRO NA BATIDA
            try:
                negativo = "-" in str(valor_hrtot) or valor_hrtot.startswith("−")
            except:
                negativo = False

            if negativo:
                for cel in linha_celulas:
                    cel.fill = preenchimento_azul

                aba_ocorrencias.append([nome_aba,data,"Erro na batida do ponto",""])

    # Ajusta formatação da aba OCORRÊNCIAS
    for coluna in aba_ocorrencias.columns:
        coluna_letra = coluna[0].column_letter
        max_len = max(len(str(cel.value)) if cel.value else 0 for cel in coluna)
        aba_ocorrencias.column_dimensions[coluna_letra].width = max_len + 2

    for linha in aba_ocorrencias.iter_rows(min_row=2):
        for cel in linha:
            cel.alignment = center_align

    wb.save(caminho_arquivo)
    print(f"Arquivo salvo com sucesso:\n{caminho_arquivo}")
