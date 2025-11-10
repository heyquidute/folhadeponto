import pdfplumber
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os

def gerar_excel(pdf_path, output_path="folha_ponto_processada.xlsx", progress_callback = None, cancel_flag = None):
    """
    Lê o PDF de folha de ponto e gera um arquivo Excel com uma aba por funcionário.
    - pdf_path: caminho do arquivo PDF
    - output_path: caminho de saída do Excel
    - progress_callback: função para atualizar a barra de progresso (opcional)
    - cancel_flag: função que retorna True se o processo for cancelado (opcional)
    """
    # ===== Cabeçalho padrão =====
    colunas = [
        "Data", "Dia", "Hr Ent M", "Hr Sai M", "Hr Ent T", "Hr Sai T",
        "Hr Tot T", "Hr Falta", "Hr Extra", "Hr Usada", "Ocorrencia"
    ]

    # ===== Função que normaliza uma linha de ponto =====
    def parse_line_to_11_fields(linha):
        s = linha.strip()
        if not s:
            return [""] * 11
        s = re.sub(r"^(\d{2}/\d{2}/202)(\s|$)", r"\g<1>5\2", s)
        tokens = re.split(r"\s+", s)
        if not tokens or not re.match(r"^\d{2}/\d{2}/\d{4}$", tokens[0]):
            return [""] * 10 + [s]
        data = tokens[0]
        dia = tokens[1] if len(tokens) > 1 else ""
        horas = []
        ocorrencia = ""
        i = 2
        while i < len(tokens):
            t = tokens[i]
            if re.match(r"^\d{2}:\d{2}$", t):
                horas.append(t)
            else:
                ocorrencia = " ".join(tokens[i:]).strip()
                break
            i += 1
        while len(horas) < 8:
            horas.append("")
        return [data, dia] + horas[:8] + [ocorrencia]

    # ===== Cria o arquivo Excel =====
    wb = Workbook()
    wb.remove(wb.active)  # remove a aba padrão "Sheet"

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)

        for page_num, page in enumerate(pdf.pages, start=1):
            # Checa cancelamento
            if cancel_flag and cancel_flag():
                if progress_callback:
                    progress_callback(0, "Processamento cancelado pelo usuário.")
                return None
            
            texto = page.extract_text()
            if not texto:
                continue

            page_lines = texto.split("\n")
            linhas_ponto = []
            funcionario = None

            for ln in page_lines:
                ln = ln.strip()

                # Captura linha de ponto
                if re.match(r"^\d{2}/\d{2}/\d{3,4}", ln):
                    ln = re.sub(r"(\d{2}/\d{2}/202)(\s|$)", r"\g<1>5\2", ln)
                    linhas_ponto.append(ln)
                    continue

                # Captura linha do funcionário
                m = re.match(r"^(\d{2})\s+(\d{6})\s+(.+)$", ln)
                if m:
                    funcionario = m.groups()  # (fl_reg, matricula, nome)
                    continue
            
            # ignora páginas sem funcionário
            if not funcionario:
                continue

            fl_reg, matricula, nome = funcionario
            nome_aba = nome.split("(")[0].strip()[:31]  # limita a 31 caracteres (limite do Excel)

            # Cria a aba do funcionário
            ws = wb.create_sheet(title=nome_aba)

            # ===== Cabeçalho =====
            for col_num, col_name in enumerate(colunas, start=1):
                c = ws.cell(row=1, column=col_num, value=col_name)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                c.alignment = Alignment(horizontal="center", vertical="center")

            # ===== Adiciona as linhas de ponto =====
            for i, linha_ponto in enumerate(linhas_ponto, start=2):
                dados = parse_line_to_11_fields(linha_ponto)
                for j, valor in enumerate(dados, start=1):
                    c = ws.cell(row=i, column=j, value=valor)
                    c.alignment = Alignment(horizontal="center", vertical="center")

            # ===== Adiciona a linha de identificação do funcionário =====
            linha_info = [""] * 11
            linha_info[0] = fl_reg      # coluna A -> Fl Reg
            linha_info[2] = matricula   # coluna C -> Matrícula
            linha_info[4] = nome        # coluna E -> Nome completo

            ultima_linha = len(linhas_ponto) + 2
            for j, valor in enumerate(linha_info, start=1):
                c = ws.cell(row=ultima_linha, column=j, value=valor)
                c.alignment = Alignment(horizontal="center", vertical="center")
                if valor:
                    c.font = Font(bold=True, color="000000")

            # ===== Ajusta largura das colunas =====
            for col in ws.columns:
                max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max(10, min(max_len + 2, 40))

            # ===== Bordas =====
            thin = Side(border_style="thin", color="808080")
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=11):
                for cell in row:
                    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

            # Atualiza progresso
            if progress_callback:
                progress_callback(page_num / total_pages * 100, f"Processando página: {page_num} de {total_pages}")

    # ===== Salva o arquivo final =====
    wb.save(output_path)
    if progress_callback:
        progress_callback(100, f"Concluído: {os.path.basename(output_path)} salvo com sucesso.")
    return output_path
