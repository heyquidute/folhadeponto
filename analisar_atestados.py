from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

def analisar_atestados(caminho_arquivo):
    # Carrega o workbook existente
    wb = load_workbook(caminho_arquivo)

    # Remove a aba "ATESTADOS" se já existir
    if "ATESTADOS" in wb.sheetnames:
        del wb["ATESTADOS"]

    # Cria nova aba ATESTADOS no início
    aba_atestados = wb.create_sheet("ATESTADOS", 0)
    aba_atestados.sheet_view.showGridLines = False

    # Estilos
    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    borda_inferior = Border(bottom=Side(style="thin", color="000000"))
    preenchimento_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    preenchimento_amarelo = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Cabeçalho da aba ATESTADOS
    cabecalho = ["Funcionário", "Data", "Detalhe","Observação"]
    aba_atestados.append(cabecalho)
    for cel in aba_atestados[1]:
        cel.font = bold_font
        cel.alignment = center_align
        cel.border = borda_inferior

    # Lista das abas dos funcionários com atestados
    funcionarios_com_atestado = set()

    # Percorre cada aba de funcionário
    for nome_aba in wb.sheetnames.copy():
        if nome_aba == "ATESTADOS":
            continue

        aba = wb[nome_aba]

        # Pega cabeçalhos da aba
        cabecalhos = [celula.value for celula in next(aba.iter_rows(min_row=1, max_row=1))]
        if "Data" not in cabecalhos or "Ocorrencia" not in cabecalhos or "Total" not in cabecalhos:
            continue

        idx_data = cabecalhos.index("Data") + 1
        idx_ocorrencia = cabecalhos.index("Ocorrencia") + 1
        idx_total = cabecalhos.index("Total") + 1

        # Percorre as linhas da aba do funcionário
        for linha_celulas in aba.iter_rows(min_row=2):
            data = linha_celulas[idx_data - 1].value
            ocorrencia_raw = linha_celulas[idx_ocorrencia - 1].value
            total_raw = linha_celulas[idx_total - 1].value
            if not ocorrencia_raw:
                continue

            ocorrencia = str(ocorrencia_raw).strip().upper()

            # Se for atestado
            if (
                ocorrencia.startswith("007") or 
                ocorrencia.startswith("ATESTADO") or 
                ocorrencia.startswith("008") or 
                ocorrencia.startswith("004")
            ):
                # Adiciona à aba ATESTADOS
                aba_atestados.append([nome_aba, data, ocorrencia_raw, ""])
                funcionarios_com_atestado.add(nome_aba)
                # Pinta a linha inteira de verde na aba do funcionário
                for cel in linha_celulas:
                    cel.fill = preenchimento_verde

            elif ocorrencia.startswith("014"):
                horas = None
                if isinstance(total_raw, timedelta):
                    horas = total_raw.total_seconds() / 3600
                elif isinstance(total_raw, datetime):
                    horas = total_raw.hour + total_raw.minute / 60
                elif isinstance(total_raw, str):
                    try:
                        partes = total_raw.split(":")
                        horas = int(partes[0]) + int(partes[1]) / 60
                    except:
                        pass
                if horas is not None and horas < 5:
                    horas_compensar = 6 - horas
                    h, m = divmod(horas_compensar * 60, 60)
                    aba_atestados.append([nome_aba, data, ocorrencia_raw, f"{int(h):02d}h{int(m):02d}min a compensar"])
                    funcionarios_com_atestado.add(nome_aba)

                    # Pinta a linha de amarelo
                    for cel in linha_celulas:
                        cel.fill = preenchimento_amarelo

    # Remove aba de quem não tem atestado
    for nome_aba in wb.sheetnames.copy():
        if nome_aba != "ATESTADOS" and nome_aba not in funcionarios_com_atestado:
            del wb[nome_aba]
    

    # Ajusta formatação da aba ATESTADOS
    for coluna in aba_atestados.columns:
        coluna_letra = coluna[0].column_letter
        max_len = max(len(str(cel.value)) if cel.value else 0 for cel in coluna)
        aba_atestados.column_dimensions[coluna_letra].width = max_len + 2

    for linha in aba_atestados.iter_rows(min_row=2):
        for cel in linha:
            cel.alignment = center_align

    wb.save(caminho_arquivo)
    print(f"Arquivo salvo com sucesso:\n{caminho_arquivo}")
