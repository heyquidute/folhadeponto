from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from convert import to_time, to_float
from cria_link import link_aba_funcionario, link_retorno

def analisar_conformidade(caminho_arquivo):
    # Carrega o workbook existente
    wb = load_workbook(caminho_arquivo)

    # Remove a aba "RESUMO" se já existir
    if "RESUMO" in wb.sheetnames:
        del wb["RESUMO"]

    # Cria nova aba RESUMO no início
    aba_resumo = wb.create_sheet("RESUMO", 0)
    aba_resumo.sheet_view.showGridLines = False

    # Estilos
    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    borda_inferior = Border(bottom=Side(style="thin", color="000000"))
    preenchimento_amarelo = PatternFill(start_color="FFF467", end_color="FFF467", fill_type="solid")
    preenchimento_laranja = PatternFill(start_color="FAA441", end_color="FAA441", fill_type="solid")
    preenchimento_azul = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Cabeçalho da aba RESUMO
    cabecalho = ["Funcionário", "Data", "Ocorrência","Informações da folha de ponto"]
    aba_resumo.append(cabecalho)
    for cel in aba_resumo[1]:
        cel.font = bold_font
        cel.alignment = center_align
        cel.border = borda_inferior

    # Percorre cada aba de funcionário
    for nome_aba in wb.sheetnames.copy():
        if nome_aba == "RESUMO":
            continue

        aba = wb[nome_aba]

        # Pega cabeçalhos da aba
        cabecalhos = [celula.value for celula in next(aba.iter_rows(min_row=1, max_row=1))]
        if "Data" not in cabecalhos or "Ocorrencia" not in cabecalhos or "Total" not in cabecalhos:
            continue

        idx_data = cabecalhos.index("Data") + 1
        idx_hrtot = cabecalhos.index("Hr Tot T")+1

        max_col = aba.max_column
        max_linha = aba.max_row
        col_turno_manha = max_col - 3
        col_t_almoco = max_col - 2
        col_turno_tarde = max_col - 1
        col_total = max_col

        # Percorre as linhas da aba do funcionário (menos a última)
        for linha_celulas in aba.iter_rows(min_row=2, max_row=max_linha-1):
            data = linha_celulas[idx_data - 1].value
            t_manha = to_time(linha_celulas[col_turno_manha - 1].value)
            t_almoco = to_time(linha_celulas[col_t_almoco - 1].value)
            t_tarde = to_time(linha_celulas[col_turno_tarde - 1].value)
            t_total = to_time(linha_celulas[col_total - 1].value)
            valor_hrtot = linha_celulas[idx_hrtot-1].value

            # TEMPO DE ALMOÇO MENOR QUE 1 HORA
            if t_almoco and t_almoco < timedelta(hours=1):
                for cell in linha_celulas:
                    cell.fill = preenchimento_amarelo
                aba_resumo.append([nome_aba, data, "Almoço < 1h", f"Tempo de almoço: {t_almoco}"])

                # cria hyperlink para aba da ocorrencia
                link_aba_funcionario(aba_resumo=aba_resumo, linha_celulas=linha_celulas, nome_aba=nome_aba, coluna="M")
            
            # TEMPO DE ALMOÇO MAIOR QUE 1 HORA E 20 MINUTOS
            elif t_almoco and t_almoco > timedelta(hours=1, minutes=20):
                for cell in linha_celulas:
                    cell.fill = preenchimento_amarelo
                aba_resumo.append([nome_aba, data, "Almoço > 1h20min", f"Tempo de almoço: {t_almoco}"])

                # cria hyperlink para aba da ocorrencia
                link_aba_funcionario(aba_resumo=aba_resumo, linha_celulas=linha_celulas, nome_aba=nome_aba, coluna="M")
                
            # TURNO COM MAIS DE 6 HORAS SEM INTERVALO
            if t_manha and t_manha > timedelta(hours=6):
                for cell in linha_celulas:
                    cell.fill = preenchimento_amarelo
                aba_resumo.append([nome_aba, data, "Turno da manhã > 6h", f"Duração do turno: {t_manha}"])

                # cria hyperlink para aba da ocorrencia
                link_aba_funcionario(aba_resumo=aba_resumo, linha_celulas=linha_celulas, nome_aba=nome_aba, coluna="L")

            if t_tarde and t_tarde > timedelta(hours=6):
                for cell in linha_celulas:
                    cell.fill = preenchimento_amarelo
                aba_resumo.append([nome_aba, data, "Turno da tarde > 6h", f"Duração do turno: {t_tarde}"])

                # cria hyperlink para aba da ocorrencia
                link_aba_funcionario(aba_resumo=aba_resumo, linha_celulas=linha_celulas, nome_aba=nome_aba, coluna="N")
            
            # JORNADA TOTAL ACIMA DE 10 HORAS
            if t_total and t_total > timedelta(hours=10):
                for cell in linha_celulas:
                    cell.fill = preenchimento_amarelo
                aba_resumo.append([nome_aba, data, "Jornada > 10h", f"Jornada total: {t_total}"])

                # cria hyperlink para aba da ocorrencia
                link_aba_funcionario(aba_resumo=aba_resumo, linha_celulas=linha_celulas, nome_aba=nome_aba, coluna="O")

            #ERRO NA BATIDA
            try:
                negativo = "-" in str(valor_hrtot) or valor_hrtot.startswith("−")
            except:
                negativo = False

            if negativo:
                for cel in linha_celulas:
                    cel.fill = preenchimento_laranja

                aba_resumo.append([nome_aba,data,"Diferente de 4 batidas",""])

                # cria hyperlink para aba da ocorrencia
                link_aba_funcionario(aba_resumo=aba_resumo, linha_celulas=linha_celulas, nome_aba=nome_aba, coluna="G")
        
        #SALDO DE HORAS NEGATIVO
        valor_saldo_hora = to_float(aba.cell(row=max_linha, column=9).value)
        if valor_saldo_hora < 0:
            for cel in aba[max_linha]:
                cel.fill = preenchimento_azul
            aba_resumo.append([nome_aba,"", "Saldo de hora negativo", f"Saldo atual: {valor_saldo_hora}"])

            # cria hyperlink para aba da ocorrencia
            link_aba_funcionario(
                aba_resumo=aba_resumo, 
                linha_celulas=[aba.cell(row=max_linha, column=1)], 
                nome_aba=nome_aba, 
                coluna="I"
            )

    # Ajusta formatação da aba RESUMO
    for coluna in aba_resumo.columns:
        coluna_letra = coluna[0].column_letter
        max_len = max(len(str(cel.value)) if cel.value else 0 for cel in coluna)
        aba_resumo.column_dimensions[coluna_letra].width = max_len + 2

    for linha in aba_resumo.iter_rows(min_row=2):
        for cel in linha:
            cel.alignment = center_align

    link_retorno(wb)
    wb.save(caminho_arquivo)
    print(f"Arquivo salvo com sucesso:\n{caminho_arquivo}")
