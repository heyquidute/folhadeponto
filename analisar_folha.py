import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from datetime import timedelta
import re
import os

# Estilos
preenchimento_amarelo = PatternFill(start_color="F9E700", end_color="F9E700", fill_type="solid")
preenchimento_laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
preenchimento_vermelho = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
bold_font = Font(bold=True)
borda_inferior = Border(bottom=Side(style="thin", color="000000"))

def str_para_tempo(valor):
    # Converte string "HH:MM" para objeto datetime.time
    if not valor:
        return None
    try:
        h, m = map(int, re.findall(r'\d+', valor))
        return timedelta(hours=h, minutes=m)
    except:
        return None   
    
def analisar_folha(caminho_excel):
    print(f"Analisando arquivo: {os.path.basename(caminho_excel)}")

    wb = openpyxl.load_workbook(caminho_excel)

    # Remove a aba RESUMO se já existir
    if "RESUMO" in wb.sheetnames:
        del wb["RESUMO"]
    
    # Cria nova aba RESUMO e move para o início
    resumo = wb.create_sheet("RESUMO")
    wb.move_sheet(resumo, offset=-(len(wb.sheetnames)-1))
    resumo.sheet_view.showGridLines = False  # 🔹 Oculta linhas de grade

    # Cabeçalho
    resumo.append(["Funcionário", "Data", "Tipo de Erro", "Detalhes"])
    for cell in resumo[1]:
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = borda_inferior

    # --- Central da análise ---
    for nome_aba in wb.sheetnames:
        if nome_aba == "RESUMO":
            continue

        ws = wb[nome_aba]

        max_col = ws.max_column
        col_turno_manha = max_col - 3
        col_t_almoco = max_col - 2
        col_turno_tarde = max_col - 1
        col_total = max_col

        for row in ws.iter_rows(min_row=2):
            try:
                data = row[0].value
                t_manha = str_para_tempo(row[col_turno_manha - 1].value)
                t_almoco = str_para_tempo(row[col_t_almoco - 1].value)
                t_tarde = str_para_tempo(row[col_turno_tarde - 1].value)
                t_total = str_para_tempo(row[col_total - 1].value)

                # Jornada total acima de 10 horas
                if t_total and t_total > timedelta(hours=10):
                    for cell in row:
                        cell.fill = preenchimento_vermelho
                    resumo.append([nome_aba, data, "Jornada > 10h", f"Jornada total: {t_total}"])

                # Turno com mais de 6 horas sem intervalo
                if t_manha and t_manha > timedelta(hours=6):
                    for cell in row:
                        cell.fill = preenchimento_laranja
                    resumo.append([nome_aba, data, "Turno da manhã > 6h", f"Duração do turno: {t_manha}"])

                if t_tarde and t_tarde > timedelta(hours=6):
                    for cell in row:
                        cell.fill = preenchimento_amarelo
                    resumo.append([nome_aba, data, "Turno da tarde > 6h", f"Duração do turno: {t_tarde}"])

                # Tempo do almoço menor que 1 hora
                if t_almoco and t_almoco < timedelta(hours=1):
                    for cell in row:
                        cell.fill = fill_amarelo
                    resumo.append([nome_aba, data, "Almoço < 1h", f"Tempo de almoço: {t_almoco}"])

            except Exception as e:
                print(f"Erro ao analisar linha {row[0].row} na aba {nome_aba}: {e}")
                continue
        
    # Centraliza todo o conteúdo do RESUMO
    for linha in resumo.iter_rows(min_row=2):
        for cell in linha:
            cell.alignment = center_align

    # Ajusta largura das colunas conforme o conteúdo
    for coluna in resumo.columns:
        max_length = 0
        coluna_letra = coluna[0].column_letter
        for cell in coluna:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        resumo.column_dimensions[coluna_letra].width = max_length + 2

    wb.save(caminho_excel)
    print("Analise concluida e arquivo salvo com sucesso.")
