from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from cria_link import link_aba_funcionario, link_retorno

def analisar_verificacao(caminho_arquivo):
    # Carrega o workbook existente
    wb = load_workbook(caminho_arquivo)

    # Remove a aba "RESUMO" se já existir
    if "RESUMO" in wb.sheetnames:
        del wb["RESUMO"]

    # Cria nova aba RESUMO no início
    aba_resumo = wb.create_sheet("RESUMO", 0)
    aba_resumo.sheet_view.showGridLines = False

    # Estilos
    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    borda_inferior = Border(bottom=Side(style="thin", color="000000"))
    preenchimento_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Cabeçalho da aba RESUMO
    cabecalho = ["Funcionário", "Data", "Ocorrência","Observações da folha de ponto"]
    aba_resumo.append(cabecalho)
    for cel in aba_resumo[1]:
        cel.font = bold_font
        cel.alignment = center_align
        cel.border = borda_inferior

    # Lista das abas dos funcionários com atestados
    funcionarios_com_atestado = set()

    # Percorre cada aba de funcionário
    for nome_aba in wb.sheetnames.copy():
        if nome_aba == "RESUMO":
            continue

        aba = wb[nome_aba]

        # Pega cabeçalhos da aba
        cabecalhos = [celula.value for celula in next(aba.iter_rows(min_row=1, max_row=1))]
        if "Data" not in cabecalhos or "Ocorrência" not in cabecalhos or "Total" not in cabecalhos:
            continue

        idx_data = cabecalhos.index("Data") + 1
        idx_ocorrencia = cabecalhos.index("Ocorrência") + 1
        idx_total = cabecalhos.index("Total") + 1
        idx_falta = cabecalhos.index("Hr Falta") + 1

        # Percorre as linhas da aba do funcionário
        for linha_celulas in aba.iter_rows(min_row=2):
            data = linha_celulas[idx_data - 1].value
            ocorrencia_raw = linha_celulas[idx_ocorrencia - 1].value
            hr_falta = linha_celulas[idx_falta - 1].value

            ocorrencia = str(ocorrencia_raw).strip().upper()

            # ATESTADOS MÉDICOS
            if (ocorrencia.startswith("007") or ocorrencia.startswith("ATESTADO")):
                aba_resumo.append([nome_aba, data, "Atestado médico", ocorrencia_raw])
                funcionarios_com_atestado.add(nome_aba)
                # Pinta a linha
                for cel in linha_celulas:
                    cel.fill = preenchimento_verde
                link_aba_funcionario(
                    aba_resumo=aba_resumo, 
                    linha_celulas=linha_celulas, 
                    nome_aba=nome_aba, 
                    coluna="K"
                    )
            
            # COMPENSAÇÃO DE HORAS
            elif ocorrencia.startswith("434"):
                aba_resumo.append([nome_aba, data, "Compensação de horas", ocorrencia_raw])
                # Pinta a linha
                for cel in linha_celulas:
                    cel.fill = preenchimento_verde
                link_aba_funcionario(
                    aba_resumo=aba_resumo,
                    linha_celulas=linha_celulas,
                    nome_aba=nome_aba,
                    coluna="K"
                )

            # SUSPENSÃO
            elif ocorrencia.startswith("010") or ocorrencia.startswith("SUSPENS"):
                aba_resumo.append([nome_aba, data, "Suspensão", ocorrencia_raw])
                # Pinta a linha
                for cel in linha_celulas:
                    cel.fill = preenchimento_verde
                link_aba_funcionario(
                    aba_resumo=aba_resumo,
                    linha_celulas=linha_celulas,
                    nome_aba=nome_aba,
                    coluna="K"
                )

            # BANCO DE HORAS DEVENDO
            elif ocorrencia.startswith("008") or ocorrencia.startswith("BANCO DE HORA"):
                aba_resumo.append([nome_aba, data, "Banco de horas", ocorrencia_raw])
                funcionarios_com_atestado.add(nome_aba)
                # Pinta a linha
                for cel in linha_celulas:
                    cel.fill = preenchimento_verde
                link_aba_funcionario(aba_resumo=aba_resumo, linha_celulas=linha_celulas, nome_aba=nome_aba, coluna="K")
                
            # ABONO
            elif ocorrencia.startswith("004") or ocorrencia.startswith("ABONO"):
                aba_resumo.append([nome_aba, data, "Abono", ocorrencia_raw])
                funcionarios_com_atestado.add(nome_aba)
                # Pinta a linha
                for cel in linha_celulas:
                    cel.fill = preenchimento_verde
                link_aba_funcionario(aba_resumo=aba_resumo, linha_celulas=linha_celulas, nome_aba=nome_aba, coluna="K")

            # SAÍDA ANTECIPADA
            elif ocorrencia.startswith("014"):
                aba_resumo.append([nome_aba, data, "Saída antecipada", f"{ocorrencia_raw} - {hr_falta}"])
                link_aba_funcionario(
                    aba_resumo=aba_resumo,
                    linha_celulas=linha_celulas,
                    nome_aba=nome_aba,
                    coluna="K"
                )
                for cel in linha_celulas:
                    cel.fill = preenchimento_verde


    # Ajusta formatação da aba RESUMO
    for coluna in aba_resumo.columns:
        coluna_letra = coluna[0].column_letter
        max_len = max(len(str(cel.value)) if cel.value else 0 for cel in coluna)
        aba_resumo.column_dimensions[coluna_letra].width = max_len + 2

    for linha in aba_resumo.iter_rows(min_row=2):
        for cel in linha:
            cel.alignment = center_align

    link_retorno(wb)
    wb.save(caminho_arquivo)
    print(f"Arquivo salvo com sucesso:\n{caminho_arquivo}")
