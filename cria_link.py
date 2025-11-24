from openpyxl.utils import quote_sheetname

def link_aba_funcionario(aba_resumo, linha_celulas, nome_aba, coluna):
    # cria hyperlink para aba da ocorrencia
        linha_resumo = aba_resumo.max_row
        linha_destino = linha_celulas[0].row
        celula_link = aba_resumo.cell(row=linha_resumo, column=1) 

        celula_link.hyperlink = f"#'{nome_aba}'!{coluna}{linha_destino+1}"


def link_retorno(wb):
    # cria hyperlink para aba da RESUMO
    nome_resumo = "RESUMO"
    nome_seguro = quote_sheetname(nome_resumo)

    for ws in wb.worksheets:
        if ws.title == nome_resumo:
            continue

        ws.insert_rows(1)

        # coloca o link na célula A1, por exemplo
        cell = ws.cell(row=1, column=1)
        cell.value = "← RESUMO"
        cell.hyperlink = f"#{nome_seguro}!A1"
        cell.style = "Hyperlink"
